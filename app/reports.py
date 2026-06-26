from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import Group, FundRound, PaymentDue, Expense
from .services import money

REPORT_DIR = Path("reports")
REPORT_DIR.mkdir(exist_ok=True)


def create_excel_report(db: Session, group: Group, round_id: int | None = None) -> Path:
    if round_id is None:
        r = db.scalar(select(FundRound).where(FundRound.group_id == group.id, FundRound.is_open == True).order_by(FundRound.created_at.desc()).limit(1))
    else:
        r = db.get(FundRound, round_id)
    if not r:
        raise ValueError("ยังไม่มีรอบเดือน")
    dues = db.scalars(select(PaymentDue).where(PaymentDue.round_id == r.id)).all()
    expenses = db.scalars(select(Expense).where(Expense.round_id == r.id)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "รายงานกองกลาง"
    ws.merge_cells("A1:H1")
    ws["A1"] = group.name or "รายงานเงินกองกลาง"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"ประจำเดือน {r.month_label}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["รายรับ", "ยอดต้องชำระ", "ยอดชำระแล้ว", "สถานะ", "วันที่", "รายจ่าย", "จำนวนเงิน", "หมายเหตุ"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h); c.font = Font(bold=True); c.alignment = Alignment(horizontal="center")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_rows = max(len(dues), len(expenses), 1)
    for i in range(max_rows):
        row = 5 + i
        if i < len(dues):
            d = dues[i]
            ws.cell(row, 1, d.member.display_name)
            ws.cell(row, 2, money(d.amount_due))
            ws.cell(row, 3, money(d.amount_paid))
            ws.cell(row, 4, d.status)
        if i < len(expenses):
            e = expenses[i]
            ws.cell(row, 5, e.item_date or "")
            ws.cell(row, 6, e.description)
            ws.cell(row, 7, money(e.amount))
        for col in range(1, 9):
            ws.cell(row, col).border = border

    total_row = 6 + max_rows
    total_paid = sum(money(d.amount_paid) for d in dues)
    total_due = sum(money(d.amount_due) for d in dues)
    total_exp = sum(money(e.amount) for e in expenses)
    balance = money(r.opening_balance) + total_paid - total_exp
    ws.cell(total_row, 1, "รวมรายรับ").font = Font(bold=True)
    ws.cell(total_row, 3, total_paid).font = Font(bold=True)
    ws.cell(total_row, 6, "รวมรายจ่าย").font = Font(bold=True)
    ws.cell(total_row, 7, total_exp).font = Font(bold=True)
    ws.cell(total_row+2, 6, "ยอดยกมา").font = Font(bold=True)
    ws.cell(total_row+2, 7, money(r.opening_balance))
    ws.cell(total_row+3, 6, "เงินคงเหลือ").font = Font(bold=True)
    ws.cell(total_row+3, 7, balance).font = Font(bold=True)

    for col in [2,3,7]:
        for row in range(5, total_row+4):
            ws.cell(row, col).number_format = '#,##0.00'
    widths = {1:24, 2:14, 3:14, 4:14, 5:14, 6:35, 7:14, 8:20}
    for col, width in widths.items():
        ws.column_dimensions[chr(64+col)].width = width
    filename = f"fund_report_{group.id}_{r.month_label}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.xlsx".replace(" ", "_")
    path = REPORT_DIR / filename
    wb.save(path)
    return path
