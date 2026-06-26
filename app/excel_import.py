from openpyxl import load_workbook
from .db import upsert_member, add_expense


def import_members_from_excel(path: str, group_id: str, sheet_name: str | None = None) -> int:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    count = 0
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        amount = ws.cell(row=row, column=2).value
        if not name or amount is None:
            continue
        name_s = str(name).strip()
        if any(x in name_s for x in ["ยอดยกมา", "รวมรายรับ", "รายรับ", "สำนักงาน"]):
            continue
        if isinstance(amount, (int, float)) and amount > 0:
            upsert_member(group_id, name_s, float(amount))
            count += 1
    return count
